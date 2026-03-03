#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
品类选品分析统一报告生成器
支持: Markdown, Excel, HTML, CSV 四种格式
所有报告保存到统一的输出文件夹
"""

import os
import sys
import json
import csv
import shutil
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional


class CategoryReportGenerator:
    """品类选品分析报告生成器"""

    def __init__(self, data: Dict, output_dir: Optional[str] = None):
        """
        初始化报告生成器

        Args:
            data: 包含 statistics, products, scores 的字典
            output_dir: 输出目录，默认为 category-reports/{品类名}_{日期}/
        """
        self.data = data
        self.statistics = data.get('statistics', {})
        self.products = data.get('products', [])
        self.scores = data.get('scores', {})

        # 确定输出目录
        if output_dir:
            self.output_dir = Path(output_dir)
        else:
            # 使用品类名和日期创建文件夹
            category_name = self._get_category_name()
            date_str = datetime.now().strftime('%Y%m%d')
            safe_name = self._sanitize_filename(category_name)
            self.output_dir = Path('category-reports') / f"{safe_name}_{date_str}"

        # 创建输出目录
        self.output_dir.mkdir(parents=True, exist_ok=True)

        # 创建数据子目录
        self.data_dir = self.output_dir / 'data'
        self.data_dir.mkdir(exist_ok=True)

        self.generated_files = []

    def _get_category_name(self) -> str:
        """从数据中获取品类名称"""
        # 尝试从各个位置获取品类名
        if 'category_name' in self.data:
            return self.data['category_name']
        if '品类名称' in self.data:
            return self.data['品类名称']
        return 'Unknown_Category'

    def _sanitize_filename(self, name: str) -> str:
        """清理文件名，移除非法字符"""
        # 移除或替换非法字符
        illegal_chars = '<>:"/\\|?*'
        for char in illegal_chars:
            name = name.replace(char, '_')
        # 移除空格
        name = name.replace(' ', '_')
        return name[:100]  # 限制长度

    def generate_all(self) -> Dict[str, str]:
        """
        生成所有格式的报告

        Returns:
            生成的文件路径字典 {格式: 文件路径}
        """
        results = {}

        # 1. Markdown 报告
        try:
            md_path = self.generate_markdown()
            results['markdown'] = str(md_path)
        except Exception as e:
            print(f"Markdown 生成失败: {e}")

        # 2. CSV 数据文件
        try:
            csv_path = self.generate_csv()
            results['csv'] = str(csv_path)
        except Exception as e:
            print(f"CSV 生成失败: {e}")

        # 3. Excel 报告 (需要 openpyxl)
        try:
            excel_path = self.generate_excel()
            results['excel'] = str(excel_path)
        except ImportError:
            print("Excel 生成跳过 (需要 openpyxl)")
        except Exception as e:
            print(f"Excel 生成失败: {e}")

        # 4. HTML 仪表板
        try:
            html_path = self.generate_html()
            results['html'] = str(html_path)
        except Exception as e:
            print(f"HTML 生成失败: {e}")

        # 5. 保存原始 JSON 数据
        try:
            json_path = self.generate_json()
            results['json'] = str(json_path)
        except Exception as e:
            print(f"JSON 生成失败: {e}")

        return results

    def generate_markdown(self) -> Path:
        """生成 Markdown 报告"""
        filename = self.output_dir / "category_analysis_report.md"

        # 读取模板
        template_path = Path(__file__).parent.parent / 'assets' / 'report_template.md'
        if template_path.exists():
            with open(template_path, 'r', encoding='utf-8') as f:
                content = f.read()
        else:
            content = self._get_default_markdown_template()

        # 替换变量
        content = self._replace_variables(content)

        # 写入文件
        with open(filename, 'w', encoding='utf-8') as f:
            f.write(content)

        self.generated_files.append(filename)
        print(f"✓ Markdown 报告: {filename}")
        return filename

    def generate_csv(self) -> Path:
        """生成 CSV 数据文件"""
        # 1. 统计数据 CSV
        stats_file = self.data_dir / "statistics.csv"
        with open(stats_file, 'w', newline='', encoding='utf-8-sig') as f:
            writer = csv.writer(f)
            writer.writerow(['指标', '数值'])
            for key, value in self.statistics.items():
                writer.writerow([key, value])

        # 2. 产品列表 CSV
        products_file = self.data_dir / "products.csv"
        with open(products_file, 'w', newline='', encoding='utf-8-sig') as f:
            fieldnames = ['ASIN', '标题', '品牌', '价格', '月销量', '评分', '排名']
            writer = csv.DictWriter(f, fieldnames=fieldnames)
            writer.writeheader()
            for i, p in enumerate(self.products, 1):
                writer.writerow({
                    'ASIN': p.get('ASIN', ''),
                    '标题': p.get('标题', '')[:100],
                    '品牌': p.get('品牌', ''),
                    '价格': p.get('价格', 0),
                    '月销量': p.get('月销量', 0),
                    '评分': p.get('评分', 0),
                    '排名': i
                })

        # 3. 评分数据 CSV
        scores_file = self.data_dir / "scores.csv"
        with open(scores_file, 'w', newline='', encoding='utf-8-sig') as f:
            writer = csv.writer(f)
            writer.writerow(['维度', '得分', '满分', '占比%'])
            max_scores = {'市场规模': 20, '增长潜力': 25, '竞争烈度': 20, '进入壁垒': 20, '利润空间': 15}
            total_max = sum(max_scores.values())
            for key, value in self.scores.items():
                if key in max_scores:
                    max_score = max_scores[key]
                    pct = (value / max_score * 100) if max_score > 0 else 0
                    writer.writerow([key, value, max_score, f"{pct:.1f}"])
                elif key not in ['总分', '评级']:
                    writer.writerow([key, value, '', ''])

        self.generated_files.extend([stats_file, products_file, scores_file])
        print(f"✓ CSV 数据文件: {self.data_dir}")
        return stats_file  # 返回主文件

    def generate_excel(self) -> Path:
        """生成 Excel 报告"""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font as OpenpyxlFont, Alignment, PatternFill
            from openpyxl.utils import get_column_letter
        except ImportError:
            raise ImportError("需要安装 openpyxl: pip install openpyxl")

        filename = self.output_dir / "category_analysis_report.xlsx"
        wb = Workbook()

        # 删除默认工作表
        wb.remove(wb.active)

        # 1. 概览工作表
        self._create_overview_sheet(wb, OpenpyxlFont)

        # 2. 产品列表工作表
        self._create_products_sheet(wb, OpenpyxlFont, PatternFill)

        # 3. 评分工作表
        self._create_scores_sheet(wb, OpenpyxlFont)

        wb.save(filename)
        self.generated_files.append(filename)
        print(f"✓ Excel 报告: {filename}")
        return filename

    def generate_html(self) -> Path:
        """生成 HTML 仪表板"""
        filename = self.output_dir / "dashboard.html"

        # 读取模板
        template_path = Path(__file__).parent.parent / 'assets' / 'dashboard_template.html'
        if template_path.exists():
            with open(template_path, 'r', encoding='utf-8') as f:
                content = f.read()
        else:
            content = self._get_default_html_template()

        # 替换变量
        content = self._replace_html_variables(content)

        with open(filename, 'w', encoding='utf-8') as f:
            f.write(content)

        self.generated_files.append(filename)
        print(f"✓ HTML 仪表板: {filename}")
        return filename

    def generate_json(self) -> Path:
        """保存原始 JSON 数据"""
        filename = self.data_dir / "raw_data.json"
        with open(filename, 'w', encoding='utf-8') as f:
            json.dump(self.data, f, ensure_ascii=False, indent=2)

        self.generated_files.append(filename)
        print(f"✓ JSON 数据: {filename}")
        return filename

    def _replace_variables(self, content: str) -> str:
        """替换 Markdown 模板变量"""
        # 基础信息
        content = content.replace('{{DATE}}', datetime.now().strftime('%Y-%m-%d'))
        content = content.replace('{{TIME}}', datetime.now().strftime('%H:%M:%S'))
        content = content.replace('{{CATEGORY_NAME}}', self._get_category_name())

        # 统计数据
        for key, value in self.statistics.items():
            content = content.replace(f'{{{{STAT_{key}}}}}', str(value))

        # 评分 (包括占比计算)
        max_scores = {'市场规模': 20, '增长潜力': 25, '竞争烈度': 20, '进入壁垒': 20, '利润空间': 15}
        for key, value in self.scores.items():
            placeholder = '{{' + f'SCORE_{key}' + '}}'
            content = content.replace(placeholder, str(value))
            # 添加占比
            if key in max_scores:
                pct = (value / max_scores[key] * 100) if max_scores[key] > 0 else 0
                placeholder_pct = '{{' + f'SCORE_{key}_占比' + '}}'
                content = content.replace(placeholder_pct, f'{pct:.1f}%')

        # 产品列表
        if self.products:
            products_table = "| 排名 | ASIN | 品牌 | 价格 | 月销量 | 评分 | 标题 |\n"
            products_table += "|------|------|------|------|--------|------|------|\n"
            for i, p in enumerate(self.products[:20], 1):
                title = p.get('标题', '')[:50]
                # 安全获取数值
                price = self._safe_float(p.get('价格', 0))
                sales = self._safe_int(p.get('月销量', 0))
                rating = self._safe_float(p.get('评分', 0))

                products_table += f"| {i} | {p.get('ASIN', '')} | {p.get('品牌', '')} | "
                products_table += f"${price:.2f} | {sales:,} | "
                products_table += f"{rating:.1f} | {title}... |\n"
            content = content.replace('{{PRODUCTS_TABLE}}', products_table)

        return content

    def _safe_float(self, value) -> float:
        """安全转换为浮点数"""
        try:
            return float(str(value).replace(',', '').replace('%', ''))
        except:
            return 0.0

    def _safe_int(self, value) -> int:
        """安全转换为整数"""
        try:
            return int(float(str(value).replace(',', '').replace('%', '')))
        except:
            return 0

    def _replace_html_variables(self, content: str) -> str:
        """替换 HTML 模板变量"""
        # 基础信息
        content = content.replace('{{CATEGORY_NAME}}', self._get_category_name())
        content = content.replace('{{DATE}}', datetime.now().strftime('%Y-%m-%d'))

        # JSON 数据注入
        content = content.replace('{{STATISTICS_JSON}}', json.dumps(self.statistics, ensure_ascii=False))
        content = content.replace('{{PRODUCTS_JSON}}', json.dumps(self.products[:50], ensure_ascii=False))
        content = content.replace('{{SCORES_JSON}}', json.dumps(self.scores, ensure_ascii=False))

        return content

    def _create_overview_sheet(self, wb, Font):
        """创建概览工作表"""
        ws = wb.create_sheet("概览")

        # 标题
        ws['A1'] = f"{self._get_category_name()} 品类选品分析报告"
        ws['A1'].font = Font(bold=True, size=16)
        ws.merge_cells('A1:D1')

        # 日期
        ws['A2'] = f"生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"

        row = 4

        # 五维评分
        ws[f'A{row}'] = "五维评分"
        ws[f'A{row}'].font = Font(bold=True)
        row += 1

        max_scores = {'市场规模': 20, '增长潜力': 25, '竞争烈度': 20, '进入壁垒': 20, '利润空间': 15}
        for key, max_score in max_scores.items():
            score = self.scores.get(key, 0)
            ws[f'A{row}'] = key
            ws[f'B{row}'] = score
            ws[f'C{row}'] = max_score
            ws[f'D{row}'] = f"{score/max_score*100:.1f}%"
            row += 1

        row += 1
        ws[f'A{row}'] = "总分"
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'] = self.scores.get('总分', 0)
        ws[f'C{row}'] = 100

        row += 1
        ws[f'A{row}'] = "评级"
        ws[f'B{row}'] = self.scores.get('评级', '')

    def _create_products_sheet(self, wb, Font, PatternFill):
        """创建产品列表工作表"""
        ws = wb.create_sheet("产品列表")

        # 表头
        headers = ['排名', 'ASIN', '品牌', '标题', '价格', '月销量', '评分']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

        # 数据
        for row_idx, product in enumerate(self.products, 2):
            ws.cell(row=row_idx, column=1, value=row_idx - 1)
            ws.cell(row=row_idx, column=2, value=product.get('ASIN', ''))
            ws.cell(row=row_idx, column=3, value=product.get('品牌', ''))
            ws.cell(row=row_idx, column=4, value=product.get('标题', '')[:50])
            ws.cell(row=row_idx, column=5, value=product.get('价格', 0))
            ws.cell(row=row_idx, column=6, value=product.get('月销量', 0))
            ws.cell(row=row_idx, column=7, value=product.get('评分', 0))

    def _create_scores_sheet(self, wb, Font):
        """创建评分工作表"""
        ws = wb.create_sheet("评分详情")

        # 表头
        headers = ['维度', '得分', '满分', '占比']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)

        # 数据
        max_scores = {'市场规模': 20, '增长潜力': 25, '竞争烈度': 20, '进入壁垒': 20, '利润空间': 15}
        for row_idx, (key, max_score) in enumerate(max_scores.items(), 2):
            score = self.scores.get(key, 0)
            ws.cell(row=row_idx, column=1, value=key)
            ws.cell(row=row_idx, column=2, value=score)
            ws.cell(row=row_idx, column=3, value=max_score)
            ws.cell(row=row_idx, column=4, value=f"{score/max_score*100:.1f}%")

    def _get_default_markdown_template(self) -> str:
        """获取默认 Markdown 模板"""
        return """# {{CATEGORY_NAME}} 品类选品分析报告

**生成时间**: {{DATE}} {{TIME}}
**数据来源**: Sorftime MCP

---

## 执行摘要

### 综合评级: {{SCORE_评级}}

### 五维评分

| 维度 | 得分 | 满分 |
|------|------|------|
| 市场规模 | {{SCORE_市场规模}} | 20 |
| 增长潜力 | {{SCORE_增长潜力}} | 25 |
| 竞争烈度 | {{SCORE_竞争烈度}} | 20 |
| 进入壁垒 | {{SCORE_进入壁垒}} | 20 |
| 利润空间 | {{SCORE_利润空间}} | 15 |
| **总分** | **{{SCORE_总分}}** | **100** |

---

## 市场数据

| 指标 | 数值 |
|------|------|
{% for key, value in statistics.items() %}| {{ key }} | {{ value }} |
{% endfor %}

---

## Top 产品列表

{{PRODUCTS_TABLE}}

---

## 数据说明

本报告基于 Sorftime MCP 实时数据生成，所有数据文件保存在 `data/` 目录中。

- `statistics.csv` - 统计数据
- `products.csv` - 产品列表
- `scores.csv` - 评分详情
- `raw_data.json` - 原始 JSON 数据

---

*报告自动生成于 {{DATE}}*
"""

    def _get_default_html_template(self) -> str:
        """获取默认 HTML 模板"""
        return """<!DOCTYPE html>
<html lang="zh-CN">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>{{CATEGORY_NAME}} 品类选品分析</title>
    <style>
        body { font-family: 'Segoe UI', sans-serif; background: #f5f7fa; padding: 20px; }
        .container { max-width: 1200px; margin: 0 auto; background: white; padding: 30px; border-radius: 10px; }
        h1 { color: #333; border-bottom: 2px solid #667eea; padding-bottom: 10px; }
        .score-card { background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); color: white; padding: 20px; border-radius: 8px; margin: 20px 0; }
        .score-item { display: flex; justify-content: space-between; padding: 5px 0; }
        table { width: 100%; border-collapse: collapse; margin: 20px 0; }
        th, td { padding: 12px; text-align: left; border-bottom: 1px solid #ddd; }
        th { background: #f8f9fa; font-weight: bold; }
    </style>
</head>
<body>
    <div class="container">
        <h1>{{CATEGORY_NAME}} 品类选品分析报告</h1>
        <p>生成时间: {{DATE}}</p>

        <div class="score-card">
            <h2>五维评分</h2>
            <div id="scores"></div>
        </div>

        <h2>产品列表</h2>
        <table id="products-table">
            <thead>
                <tr>
                    <th>排名</th>
                    <th>ASIN</th>
                    <th>品牌</th>
                    <th>价格</th>
                    <th>月销量</th>
                    <th>评分</th>
                </tr>
            </thead>
            <tbody></tbody>
        </table>
    </div>

    <script>
        const scores = {{SCORES_JSON}};
        const products = {{PRODUCTS_JSON}};

        // 渲染评分
        const scoresDiv = document.getElementById('scores');
        const maxScores = {'市场规模': 20, '增长潜力': 25, '竞争烈度': 20, '进入壁垒': 20, '利润空间': 15};
        for (const [key, score] of Object.entries(scores)) {
            if (key in maxScores) {
                const div = document.createElement('div');
                div.className = 'score-item';
                div.innerHTML = `<span>${key}</span><span>${score}/${maxScores[key]}</span>`;
                scoresDiv.appendChild(div);
            }
        }

        // 渲染产品
        const tbody = document.querySelector('#products-table tbody');
        products.slice(0, 20).forEach((p, i) => {
            const tr = document.createElement('tr');
            tr.innerHTML = `
                <td>${i + 1}</td>
                <td>${p.ASIN || ''}</td>
                <td>${p.品牌 || ''}</td>
                <td>$${(p.价格 || 0).toFixed(2)}</td>
                <td>${(p.月销量 || 0).toLocaleString()}</td>
                <td>${(p.评分 || 0).toFixed(1)}★</td>
            `;
            tbody.appendChild(tr);
        });
    </script>
</body>
</html>
"""


def main():
    """命令行入口"""
    if len(sys.argv) < 2:
        print("用法: python generate_reports.py <JSON数据文件> [输出目录]")
        print("\n示例:")
        print("  python generate_reports.py parsed_data.json")
        print("  python generate_reports.py parsed_data.json ./reports")
        sys.exit(1)

    json_file = sys.argv[1]
    output_dir = sys.argv[2] if len(sys.argv) > 2 else None

    # 读取数据
    with open(json_file, 'r', encoding='utf-8') as f:
        data = json.load(f)

    # 生成报告
    generator = CategoryReportGenerator(data, output_dir)
    results = generator.generate_all()

    print("\n" + "=" * 60)
    print("报告生成完成！")
    print("=" * 60)
    print(f"输出目录: {generator.output_dir}")
    print("\n生成的文件:")
    for format_type, path in results.items():
        print(f"  [{format_type.upper()}] {path}")


if __name__ == "__main__":
    main()
